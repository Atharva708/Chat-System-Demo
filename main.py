from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from typing import List, Dict, Optional
import json
import base64
from datetime import datetime
import asyncio
import pandas as pd
import os
import io
import platform
from extractor import extract_attributes, ConversationData, asdict
from openpyxl import Workbook, load_workbook
from io import BytesIO
import gspread
from google.oauth2.service_account import Credentials
import uuid
import pytesseract
import pypdfium2 as pdfium
from PIL import Image, ImageOps, ImageFilter
from docx import Document

app = FastAPI()

# In-memory storage for daily Excel files (for download fallback)
# Key: date string (YYYY-MM-DD), Value: (Workbook object, file_bytes)
daily_excel_files: Dict[str, tuple] = {}

# Store active WebSocket connections with user info
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[WebSocket, Dict] = {}  # websocket -> user_info
        self.messages: List[Dict] = []  # Store messages per session
    
    async def connect(self, websocket: WebSocket, user_name: str = None, user_id: str = None):
        await websocket.accept()
        # Generate user identifier
        user_identifier = user_name or user_id or f"User_{id(websocket)}"
        self.active_connections[websocket] = {
            "name": user_name or user_identifier,
            "id": user_id or user_identifier,
            "identifier": user_identifier
        }
        # Send chat history to new connection
        if self.messages:
            await websocket.send_json({
                "type": "history",
                "messages": self.messages
            })
    
    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            del self.active_connections[websocket]
    
    def get_user_info(self, websocket: WebSocket) -> Dict:
        return self.active_connections.get(websocket, {"name": "Anonymous", "id": "unknown", "identifier": "Anonymous"})
    
    async def broadcast(self, message: dict):
        # Add message to history
        self.messages.append(message)
        # Keep only last 100 messages to prevent memory issues
        if len(self.messages) > 100:
            self.messages.pop(0)
        
        # Broadcast to all connected clients
        for connection in list(self.active_connections.keys()):
            try:
                await connection.send_json(message)
            except:
                pass  # Skip failed connections

manager = ConnectionManager()

# OCR Configuration - Local Tesseract
def configure_tesseract():
    """
    Configure Tesseract OCR path.
    1) If env var TESSERACT_CMD is set, use it
    2) Else on Windows, fallback to default install path
    3) Else rely on PATH (Linux/macOS/Docker)
    """
    env_cmd = os.getenv("TESSERACT_CMD")
    if env_cmd and os.path.exists(env_cmd):
        pytesseract.pytesseract.tesseract_cmd = env_cmd
        print(f"✓ Using Tesseract from TESSERACT_CMD: {env_cmd}")
        return

    if platform.system().lower().startswith("win"):
        win_default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(win_default):
            pytesseract.pytesseract.tesseract_cmd = win_default
            print(f"✓ Using Tesseract from Windows default: {win_default}")
            return
    
    print("✓ Using Tesseract from system PATH")

configure_tesseract()

# Verify Tesseract is available and working
try:
    tesseract_version = pytesseract.get_tesseract_version()
    print(f"✓ Tesseract OCR version {tesseract_version} detected and working")
except Exception as e:
    print(f"⚠ WARNING: Tesseract OCR not found or not accessible: {e}")
    print("  Please install Tesseract:")
    print("    • macOS: brew install tesseract")
    print("    • Ubuntu/Debian: sudo apt-get install tesseract-ocr")
    print("    • Windows: Download from https://github.com/UB-Mannheim/tesseract/wiki")
    print("    • Or set TESSERACT_CMD environment variable to point to tesseract executable")

# Google Sheets Configuration
# On Render, use environment variables. Locally, try credentials.json first
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "")
GOOGLE_CREDENTIALS_JSON = os.getenv("GOOGLE_SHEETS_CREDENTIALS_JSON", "")
CREDENTIALS_FILE = "credentials.json"
IS_RENDER = os.getenv("RENDER") is not None

# On Render, prioritize environment variables. Locally, try file first
if not IS_RENDER and os.path.exists(CREDENTIALS_FILE):
    try:
        with open(CREDENTIALS_FILE, 'r') as f:
            creds_dict = json.load(f)
            # If file contains sheet_id, use it (unless env var is set)
            if 'sheet_id' in creds_dict and not GOOGLE_SHEET_ID:
                GOOGLE_SHEET_ID = creds_dict['sheet_id']
            print(f"✓ Loaded credentials from {CREDENTIALS_FILE}")
    except Exception as e:
        print(f"⚠ Error reading {CREDENTIALS_FILE}: {e}")

# Fallback: Save Excel files locally if Google Sheets not configured
DESKTOP_PATH = os.path.join(os.path.expanduser("~"), "Desktop")
EXCEL_OUTPUT_DIR = os.path.join(DESKTOP_PATH, "Chat_Extracted_Data")

# Create output directory if it doesn't exist (for local fallback)
try:
    os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)
except:
    pass  # Ignore if can't create (e.g., on Render)

# Initialize Google Sheets client if credentials are available
google_sheets_client = None
GOOGLE_SHEETS_INIT_ERROR = None

# Priority: On Render use env vars, locally try file first then env vars
if IS_RENDER:
    print("🌐 Running on Render - using environment variables for credentials")
    # On Render, must use environment variables
    if GOOGLE_CREDENTIALS_JSON and GOOGLE_SHEET_ID:
        try:
            print("📝 Initializing from environment variables...")
            creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
            scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
            google_sheets_client = gspread.authorize(creds)
            print("✓ Google Sheets client authorized from environment variables")
            
            # Verify connection
            test_sheet = google_sheets_client.open_by_key(GOOGLE_SHEET_ID)
            print(f"✓ Successfully connected to Google Sheet: {test_sheet.title}")
            print("✓ Google Sheets fully initialized and ready!")
        except json.JSONDecodeError as e:
            GOOGLE_SHEETS_INIT_ERROR = f"Invalid JSON in GOOGLE_SHEETS_CREDENTIALS_JSON: {str(e)}"
            print(f"✗ {GOOGLE_SHEETS_INIT_ERROR}")
        except Exception as e:
            GOOGLE_SHEETS_INIT_ERROR = f"Error initializing from environment variables: {str(e)}"
            print(f"✗ {GOOGLE_SHEETS_INIT_ERROR}")
            import traceback
            traceback.print_exc()
    else:
        missing = []
        if not GOOGLE_CREDENTIALS_JSON:
            missing.append("GOOGLE_SHEETS_CREDENTIALS_JSON")
        if not GOOGLE_SHEET_ID:
            missing.append("GOOGLE_SHEET_ID")
        GOOGLE_SHEETS_INIT_ERROR = f"Missing environment variables on Render: {', '.join(missing)}"
        print(f"✗ {GOOGLE_SHEETS_INIT_ERROR}")
        print("   Please set these in your Render dashboard: Environment > Add Environment Variable")
else:
    # Local development: try file first, then environment variables
    print("💻 Running locally - trying credentials.json first...")
    if os.path.exists(CREDENTIALS_FILE):
        try:
            print(f"📁 Found credentials.json at: {os.path.abspath(CREDENTIALS_FILE)}")
            with open(CREDENTIALS_FILE, 'r') as f:
                creds_dict = json.load(f)
            print("✓ Loaded credentials.json successfully")
            
            # Get sheet ID from file if present (unless env var is set)
            if 'sheet_id' in creds_dict and not GOOGLE_SHEET_ID:
                GOOGLE_SHEET_ID = creds_dict['sheet_id']
                print(f"✓ Found sheet_id in credentials.json: {GOOGLE_SHEET_ID[:20]}...")
            elif not GOOGLE_SHEET_ID:
                print("⚠ No sheet_id found in credentials.json or environment")
            
            # Initialize Google Sheets client
            scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
            google_sheets_client = gspread.authorize(creds)
            print("✓ Google Sheets client authorized")
            
            # Verify connection by trying to open the sheet
            if GOOGLE_SHEET_ID:
                try:
                    test_sheet = google_sheets_client.open_by_key(GOOGLE_SHEET_ID)
                    print(f"✓ Successfully connected to Google Sheet: {test_sheet.title}")
                    print("✓ Google Sheets fully initialized and ready!")
                except Exception as e:
                    GOOGLE_SHEETS_INIT_ERROR = f"Cannot access Google Sheet: {str(e)}. Make sure the sheet is shared with: {creds_dict.get('client_email', 'service account email')}"
                    print(f"✗ {GOOGLE_SHEETS_INIT_ERROR}")
            else:
                GOOGLE_SHEETS_INIT_ERROR = "GOOGLE_SHEET_ID not configured"
                print(f"✗ {GOOGLE_SHEETS_INIT_ERROR}")
                
        except json.JSONDecodeError as e:
            GOOGLE_SHEETS_INIT_ERROR = f"Invalid JSON in credentials.json: {str(e)}"
            print(f"✗ {GOOGLE_SHEETS_INIT_ERROR}")
        except Exception as e:
            GOOGLE_SHEETS_INIT_ERROR = f"Error initializing from credentials.json: {str(e)}"
            print(f"✗ {GOOGLE_SHEETS_INIT_ERROR}")
            import traceback
            traceback.print_exc()
    
    # Fallback to environment variables if file doesn't exist or failed
    if not google_sheets_client and GOOGLE_CREDENTIALS_JSON and GOOGLE_SHEET_ID:
        try:
            print("📝 Trying to initialize from environment variables...")
            creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
            scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
            google_sheets_client = gspread.authorize(creds)
            
            # Verify connection
            test_sheet = google_sheets_client.open_by_key(GOOGLE_SHEET_ID)
            print(f"✓ Successfully connected to Google Sheet: {test_sheet.title}")
            print("✓ Google Sheets initialized successfully from environment variables")
        except Exception as e:
            GOOGLE_SHEETS_INIT_ERROR = f"Error initializing from environment variables: {str(e)}"
            print(f"✗ {GOOGLE_SHEETS_INIT_ERROR}")
            google_sheets_client = None

# Final status
if not google_sheets_client or not GOOGLE_SHEET_ID:
    error_msg = GOOGLE_SHEETS_INIT_ERROR or "Google Sheets not configured"
    print(f"⚠ {error_msg}")
    if IS_RENDER:
        print("⚠ On Render, you MUST set these environment variables:")
        print("   1. GOOGLE_SHEETS_CREDENTIALS_JSON - Your service account JSON (entire content)")
        print("   2. GOOGLE_SHEET_ID - Your Google Sheet ID (e.g., 1wsIj3UJFlyDUaD0af-XbguRcP20La8uc0C3JP3imTgQ)")
        print("   Go to: Render Dashboard > Your Service > Environment > Add Environment Variable")

# Define all possible fields from ConversationData (for consistent column headers)
ALL_FIELDS = [
    'timestamp', 'sentiment', 'member_id', 'first_name', 'last_name', 'dob',
    'address', 'city', 'state', 'zip_code', 'address_status', 'member_status',
    'start_date', 'end_date', 'health_plan', 'contract_type', 'codes',
    'change_request', 'raw_text', 'user_identifier', 'extracted_by', 'extraction_timestamp'
]

def ocr_image_pil(image: Image.Image) -> str:
    """
    OCR for a single PIL Image using Tesseract, with robust preprocessing.
    Based on the provided OCR service code for better accuracy.
    """
    try:
        # Convert to grayscale
        image = image.convert("L")
        
        # Resize if too small (improves OCR accuracy)
        w, h = image.size
        max_side = max(w, h)
        if max_side < 1000:
            scale = 1000 / max_side
            new_w = int(w * scale)
            new_h = int(h * scale)
            image = image.resize((new_w, new_h), Image.LANCZOS)
            print(f"  Resized image from {w}x{h} to {new_w}x{new_h} for better OCR")
        
        # Enhance image for better OCR
        image = ImageOps.autocontrast(image)
        image = image.filter(ImageFilter.SHARPEN)
        
        # Perform OCR with multiple attempts if needed
        try:
            text = pytesseract.image_to_string(image, lang='eng')
            if text and text.strip():
                return text.strip()
        except Exception as e:
            print(f"  OCR with preprocessing failed: {e}, trying without preprocessing...")
        
        # Fallback: try original image
        try:
            original = image.convert("L")
            text = pytesseract.image_to_string(original, lang='eng')
            if text and text.strip():
                return text.strip()
        except Exception as e2:
            print(f"  OCR fallback failed: {e2}")
        
        return ""
    except Exception as e:
        print(f"Error in OCR preprocessing: {e}")
        import traceback
        traceback.print_exc()
        # Last resort: try direct OCR
        try:
            text = pytesseract.image_to_string(image, lang='eng')
            return text.strip() if text else ""
        except:
            return ""

async def extract_text_from_image(image_base64: str) -> Optional[str]:
    """
    Extracts text from base64-encoded image using local Tesseract OCR.
    Robust implementation with multiple fallback strategies.
    """
    try:
        # Convert base64 to bytes
        if ',' in image_base64:
            # Remove data URL prefix (e.g., "data:image/png;base64,")
            header, image_base64 = image_base64.split(',', 1)
        
        image_data = base64.b64decode(image_base64)
        print(f"📷 Processing image with local OCR ({len(image_data)} bytes)...")
        
        # Convert bytes to PIL Image
        try:
            image = Image.open(io.BytesIO(image_data))
            print(f"  Image format: {image.format}, size: {image.size}")
        except Exception as e:
            print(f"✗ Failed to open image: {e}")
            return None
        
        # Perform OCR with multiple strategies
        ocr_text = None
        
        # Strategy 1: Full preprocessing
        try:
            ocr_text = ocr_image_pil(image)
            if ocr_text and len(ocr_text.strip()) > 0:  # Accept any text
                print(f"✓ OCR successful (strategy 1), extracted {len(ocr_text)} characters")
                return ocr_text
        except Exception as e:
            print(f"  Strategy 1 failed: {e}")
        
        # Strategy 2: Direct OCR without preprocessing
        try:
            text = pytesseract.image_to_string(image, lang='eng')
            if text and text.strip() and len(text.strip()) > 0:
                print(f"✓ OCR successful (strategy 2), extracted {len(text.strip())} characters")
                return text.strip()
        except Exception as e:
            print(f"  Strategy 2 failed: {e}")
        
        # Strategy 3: Try with different image modes
        try:
            for mode in ['RGB', 'L', '1']:
                try:
                    test_image = image.convert(mode)
                    text = pytesseract.image_to_string(test_image, lang='eng')
                    if text and text.strip() and len(text.strip()) > 0:
                        print(f"✓ OCR successful (strategy 3, mode={mode}), extracted {len(text.strip())} characters")
                        return text.strip()
                except:
                    continue
        except Exception as e:
            print(f"  Strategy 3 failed: {e}")
        
        # Strategy 4: Try different PSM modes
        try:
            psm_configs = ['--psm 6', '--psm 11', '--psm 12', '--psm 3', '--psm 4']
            for config in psm_configs:
                try:
                    text = pytesseract.image_to_string(image, config=config, lang='eng')
                    if text and text.strip() and len(text.strip()) > 0:
                        print(f"✓ OCR successful (strategy 4, config={config}), extracted {len(text.strip())} characters")
                        return text.strip()
                except:
                    continue
        except Exception as e:
            print(f"  Strategy 4 failed: {e}")
        
        # If we got any text at all, return it (even if very short)
        if ocr_text and ocr_text.strip():
            print(f"⚠ OCR returned text ({len(ocr_text)} chars), returning for processing")
            return ocr_text.strip()
        
        # Last resort: try with minimal processing
        try:
            simple_text = pytesseract.image_to_string(image, lang='eng')
            if simple_text and simple_text.strip():
                print(f"⚠ OCR returned text with minimal processing ({len(simple_text.strip())} chars)")
                return simple_text.strip()
        except:
            pass
        
        print("⚠ OCR returned empty text after all strategies")
        return None
            
    except Exception as e:
        print(f"✗ Error in local OCR: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

async def process_text_locally(text: str) -> Dict:
    """
    Processes text using local extractor and returns extracted data.
    Returns ALL fields, including None values for consistent Google Sheets structure.
    """
    try:
        # Use local extractor
        extracted_data = extract_attributes(text)
        # Convert to dictionary - keep ALL fields including None values
        result_dict = asdict(extracted_data)
        # Convert None to empty string for Google Sheets compatibility
        result_dict = {k: (v if v is not None else '') for k, v in result_dict.items()}
        return {
            "status": "success",
            "extracted_data": result_dict
        }
    except Exception as e:
        print(f"Error in local extraction: {str(e)}")
        return {
            "status": "error",
            "message": str(e)
        }

def save_to_google_sheets(extracted_data: Dict, timestamp: str) -> str:
    """
    Saves extracted data to Google Sheets with ALL fields visible.
    Uses consistent column headers for all rows.
    Returns a success message with sheet name.
    """
    if not google_sheets_client:
        raise Exception("Google Sheets client not initialized. Check credentials.json and ensure the service account has access.")
    
    if not GOOGLE_SHEET_ID:
        raise Exception("GOOGLE_SHEET_ID not configured. Add 'sheet_id' to credentials.json or set GOOGLE_SHEET_ID environment variable.")
    
    try:
        # Open the spreadsheet
        print(f"📊 Opening Google Sheet with ID: {GOOGLE_SHEET_ID[:20]}...")
        sheet = google_sheets_client.open_by_key(GOOGLE_SHEET_ID)
        print(f"✓ Opened sheet: {sheet.title}")
        
        # Get today's date for worksheet name
        today = datetime.now().strftime("%Y-%m-%d")
        worksheet_name = f"Extracted Data {today}"
        
        # Try to get existing worksheet or create new one
        worksheet_exists = True
        worksheet = None
        try:
            worksheet = sheet.worksheet(worksheet_name)
            print(f"✓ Found existing worksheet: {worksheet_name}")
            # Check if headers exist
            existing_headers = worksheet.row_values(1)
            if not existing_headers or len(existing_headers) == 0:
                print("⚠ Worksheet exists but has no headers, adding headers...")
                worksheet_exists = False
            elif existing_headers != ALL_FIELDS:
                print(f"⚠ Headers don't match expected fields. Expected {len(ALL_FIELDS)} fields, found {len(existing_headers)}")
                # Headers exist but might be different - we'll still append data
        except gspread.exceptions.WorksheetNotFound:
            worksheet_exists = False
            print(f"📝 Worksheet '{worksheet_name}' not found, will create new one")
        except Exception as e:
            print(f"⚠ Error checking worksheet: {e}, will create new one")
            worksheet_exists = False
        
        if not worksheet_exists or worksheet is None:
            # Create new worksheet with enough columns
            print(f"📝 Creating new worksheet: {worksheet_name} with {len(ALL_FIELDS)} columns")
            worksheet = sheet.add_worksheet(title=worksheet_name, rows=1000, cols=len(ALL_FIELDS))
            # Add header row with ALL fields in consistent order
            worksheet.append_row(ALL_FIELDS)
            print(f"✓ Created worksheet with headers: {', '.join(ALL_FIELDS[:5])}... ({len(ALL_FIELDS)} total)")
        
        # Ensure all fields are present in extracted_data (fill missing ones with empty string)
        complete_data = {}
        for field in ALL_FIELDS:
            complete_data[field] = extracted_data.get(field, '')
        
        # Build row values in the same order as ALL_FIELDS
        row_values = [complete_data.get(field, '') for field in ALL_FIELDS]
        
        # Append data row
        print(f"💾 Appending data row to worksheet...")
        worksheet.append_row(row_values)
        
        non_empty_count = len([v for v in row_values if v])
        print(f"✓ Successfully saved to Google Sheet!")
        print(f"  Sheet: {sheet.title}")
        print(f"  Worksheet: {worksheet_name}")
        print(f"  Fields with data: {non_empty_count}/{len(ALL_FIELDS)}")
        
        return f"Google Sheet: {sheet.title} > {worksheet_name}"
        
    except gspread.exceptions.APIError as e:
        error_msg = f"Google Sheets API error: {str(e)}"
        if "PERMISSION_DENIED" in str(e) or "permission" in str(e).lower():
            error_msg += f"\nMake sure the sheet is shared with the service account email from credentials.json"
        print(f"✗ {error_msg}")
        raise Exception(error_msg)
    except Exception as e:
        error_msg = f"Error saving to Google Sheets: {str(e)}"
        print(f"✗ {error_msg}")
        import traceback
        traceback.print_exc()
        raise Exception(error_msg)

def save_to_excel_local(extracted_data: Dict, timestamp: str) -> str:
    """
    Saves extracted data to a local Excel file (fallback).
    Uses consistent column structure matching Google Sheets format.
    Returns the file path.
    """
    try:
        # Get today's date for filename
        today = datetime.now().strftime("%Y-%m-%d")
        filename = f"extracted_data_{today}.xlsx"
        filepath = os.path.join(EXCEL_OUTPUT_DIR, filename)
        
        # Ensure all fields are present (fill missing ones with empty string)
        complete_data = {}
        for field in ALL_FIELDS:
            complete_data[field] = extracted_data.get(field, '')
        
        # Build row values in the same order as ALL_FIELDS
        row_values = [complete_data.get(field, '') for field in ALL_FIELDS]
        
        # Create or append to Excel file
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "Extraction Data"
            ws.append(ALL_FIELDS)  # Header row with all fields
            ws.append(row_values)    # Data row
            wb.save(filepath)
        else:
            wb = load_workbook(filepath)
            ws = wb.active
            # Check if headers match
            existing_headers = [cell.value for cell in ws[1]]
            if existing_headers != ALL_FIELDS:
                # Headers don't match, add header row if empty
                if not existing_headers or len(existing_headers) == 0:
                    ws.insert_rows(1)
                    for col_idx, header in enumerate(ALL_FIELDS, start=1):
                        ws.cell(row=1, column=col_idx, value=header)
            ws.append(row_values)  # Append data row
            wb.save(filepath)
        
        return filepath
        
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        # Fallback: save as JSON if Excel fails
        filename = f"extracted_data_{timestamp.replace(':', '-').replace(' ', '_')}.json"
        filepath = os.path.join(EXCEL_OUTPUT_DIR, filename)
        with open(filepath, 'w') as f:
            json.dump(extracted_data, f, indent=2)
        return filepath

def append_to_daily_excel(extracted_data: Dict, timestamp: str) -> str:
    """
    Appends data to the daily Excel file in memory.
    Creates the file if it doesn't exist, or appends to existing one.
    Returns the date string (YYYY-MM-DD) used as the file key.
    """
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        
        # Ensure all fields are present
        complete_data = {}
        for field in ALL_FIELDS:
            complete_data[field] = extracted_data.get(field, '')
        
        # Build row values
        row_values = [complete_data.get(field, '') for field in ALL_FIELDS]
        
        # Get or create daily workbook
        if today not in daily_excel_files:
            # Create new workbook for today
            wb = Workbook()
            ws = wb.active
            ws.title = "Extraction Data"
            # Add headers
            ws.append(ALL_FIELDS)
            daily_excel_files[today] = (wb, None)  # None means not yet serialized
            print(f"📝 Created new daily Excel file for {today}")
        
        # Get the workbook
        wb, _ = daily_excel_files[today]
        ws = wb.active
        
        # Append data row
        ws.append(row_values)
        
        # Update stored workbook (invalidate cached bytes)
        daily_excel_files[today] = (wb, None)
        
        print(f"✓ Appended data to daily Excel file for {today} (total rows: {ws.max_row})")
        return today
    except Exception as e:
        print(f"Error appending to daily Excel: {e}")
        import traceback
        traceback.print_exc()
        raise

def get_daily_excel_bytes(date_str: str) -> bytes:
    """
    Gets the serialized bytes of the daily Excel file.
    Caches the result for performance.
    """
    if date_str not in daily_excel_files:
        raise Exception(f"No Excel file found for date: {date_str}")
    
    wb, cached_bytes = daily_excel_files[date_str]
    
    # Return cached bytes if available
    if cached_bytes is not None:
        return cached_bytes
    
    # Serialize workbook to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_bytes = excel_buffer.getvalue()
    
    # Cache the bytes
    daily_excel_files[date_str] = (wb, excel_bytes)
    
    return excel_bytes

def save_extracted_data(extracted_data: Dict, timestamp: str) -> Dict:
    """
    Saves extracted data to Google Sheets. If that fails, creates downloadable Excel.
    Returns a dict with status, message, and optional download_url.
    """
    # Always try Google Sheets first
    if google_sheets_client and GOOGLE_SHEET_ID:
        try:
            result = save_to_google_sheets(extracted_data, timestamp)
            return {
                "status": "success",
                "message": result,
                "download_url": None
            }
        except Exception as e:
            error_msg = f"Google Sheets save failed: {str(e)}"
            print(f"✗ {error_msg}")
            import traceback
            traceback.print_exc()
            # Fall through to Excel fallback
    
    # Fallback: Append to daily Excel file
    try:
        print("📥 Appending to daily Excel file as fallback...")
        date_str = append_to_daily_excel(extracted_data, timestamp)
        download_url = f"/download-daily-excel/{date_str}"
        filename = f"extracted_data_{date_str}.xlsx"
        
        return {
            "status": "excel_fallback",
            "message": f"Data saved to daily Excel file. Download at end of day.",
            "date": date_str
        }
    except Exception as e:
        error_msg = f"Failed to save to Excel file: {str(e)}"
        print(f"✗ {error_msg}")
        raise Exception(error_msg)

@app.get("/test")
async def test():
    return {"status": "Server is running!"}

@app.get("/test-sheets")
async def test_sheets():
    """Test endpoint to verify Google Sheets connection"""
    try:
        if not google_sheets_client:
            return {
                "status": "error",
                "message": "Google Sheets client not initialized",
                "details": "Check credentials.json file and ensure it's properly configured"
            }
        
        if not GOOGLE_SHEET_ID:
            return {
                "status": "error",
                "message": "GOOGLE_SHEET_ID not configured",
                "details": "Add 'sheet_id' to credentials.json or set GOOGLE_SHEET_ID environment variable"
            }
        
        # Try to open the sheet
        sheet = google_sheets_client.open_by_key(GOOGLE_SHEET_ID)
        
        # Get worksheet list
        worksheets = [ws.title for ws in sheet.worksheets()]
        
        return {
            "status": "success",
            "message": "Google Sheets connection successful!",
            "sheet_title": sheet.title,
            "sheet_id": GOOGLE_SHEET_ID,
            "worksheets": worksheets,
            "total_worksheets": len(worksheets)
        }
    except gspread.exceptions.APIError as e:
        return {
            "status": "error",
            "message": "Google Sheets API error",
            "details": str(e),
            "hint": "Make sure the sheet is shared with the service account email from credentials.json"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": "Error connecting to Google Sheets",
            "details": str(e)
        }

@app.get("/download-daily-excel/{date_str}")
async def download_daily_excel(date_str: str):
    """Download the daily Excel file for a specific date (YYYY-MM-DD format)"""
    try:
        excel_bytes = get_daily_excel_bytes(date_str)
        filename = f"extracted_data_{date_str}.xlsx"
        
        return StreamingResponse(
            BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        return {"error": f"File not found for date {date_str}: {str(e)}"}

@app.get("/download-today-excel")
async def download_today_excel():
    """Download today's Excel file (convenience endpoint)"""
    today = datetime.now().strftime("%Y-%m-%d")
    try:
        excel_bytes = get_daily_excel_bytes(today)
        filename = f"extracted_data_{today}.xlsx"
        
        return StreamingResponse(
            BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        return {"error": f"No data available for today ({today}): {str(e)}"}

@app.get("/", response_class=HTMLResponse)
async def get_homepage():
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Real-Time Chat</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            background-color: #f3f4f6;
            height: 100vh;
            overflow: hidden;
        }
        /* Custom scrollbar */
        .custom-scrollbar::-webkit-scrollbar {
            width: 8px;
        }
        .custom-scrollbar::-webkit-scrollbar-track {
            background: #f1f1f1;
        }
        .custom-scrollbar::-webkit-scrollbar-thumb {
            background: #888;
            border-radius: 4px;
        }
        .custom-scrollbar::-webkit-scrollbar-thumb:hover {
            background: #555;
        }
        /* Teams purple color */
        .teams-purple {
            background-color: #4B53BC !important;
        }
        .teams-purple-dark {
            background-color: #3d44a0 !important;
        }
        .teams-purple-light {
            background-color: #6b73d4 !important;
        }
        .teams-purple-text {
            color: #4B53BC !important;
        }
        .teams-purple-border {
            border-color: #4B53BC !important;
        }
        .teams-purple-ring:focus {
            outline: 2px solid #4B53BC;
            outline-offset: 2px;
        }
    </style>
</head>
<body class="bg-gray-100 h-screen overflow-hidden">
    <div class="flex h-screen">
        <!-- Left Sidebar -->
        <div class="w-64 bg-white border-r border-gray-200 flex flex-col">
            <div class="p-4 border-b border-gray-200">
                <h2 class="text-xl font-semibold text-gray-800">Recent Chats</h2>
            </div>
            <div class="flex-1 overflow-y-auto custom-scrollbar">
                <div class="p-2">
                    <div class="chat-item p-3 rounded-lg hover:bg-gray-100 cursor-pointer mb-2 active-chat" data-chat="Team Alpha">
                        <div class="flex items-center">
                            <div class="w-10 h-10 rounded-full teams-purple flex items-center justify-center text-white font-semibold mr-3">
                                TA
                            </div>
                            <div class="flex-1">
                                <div class="font-semibold text-gray-800">Team Alpha</div>
                                <div class="text-sm text-gray-500">Active now</div>
                            </div>
                        </div>
                    </div>
                    <div class="chat-item p-3 rounded-lg hover:bg-gray-100 cursor-pointer mb-2" data-chat="General">
                        <div class="flex items-center">
                            <div class="w-10 h-10 rounded-full bg-blue-500 flex items-center justify-center text-white font-semibold mr-3">
                                G
                            </div>
                            <div class="flex-1">
                                <div class="font-semibold text-gray-800">General</div>
                                <div class="text-sm text-gray-500">Active now</div>
                            </div>
                        </div>
                    </div>
                    <div class="chat-item p-3 rounded-lg hover:bg-gray-100 cursor-pointer mb-2" data-chat="Project Beta">
                        <div class="flex items-center">
                            <div class="w-10 h-10 rounded-full bg-green-500 flex items-center justify-center text-white font-semibold mr-3">
                                PB
                            </div>
                            <div class="flex-1">
                                <div class="font-semibold text-gray-800">Project Beta</div>
                                <div class="text-sm text-gray-500">Active now</div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- Main Chat Area -->
        <div class="flex-1 flex flex-col bg-white">
            <!-- Chat Header -->
            <div class="teams-purple text-white p-4 flex items-center justify-between border-b teams-purple-dark">
                <div class="flex items-center">
                    <div class="w-10 h-10 rounded-full teams-purple-dark flex items-center justify-center text-white font-semibold mr-3">
                        TA
                    </div>
                    <div>
                        <div class="font-semibold text-lg" id="chat-title">Team Alpha</div>
                        <div class="text-sm opacity-80">Online</div>
                    </div>
                </div>
                <div class="flex items-center space-x-2 text-sm">
                    <div class="flex items-center space-x-1">
                        <span id="ocr-status-indicator" class="w-2 h-2 rounded-full bg-green-400" title="Local OCR (Tesseract)"></span>
                        <span class="opacity-80">OCR</span>
                    </div>
                    <div class="flex items-center space-x-1">
                        <span id="extractor-status-indicator" class="w-2 h-2 rounded-full bg-green-400" title="Local Extractor"></span>
                        <span class="opacity-80">Extractor</span>
                    </div>
                    <a href="/download-today-excel" 
                       id="download-excel-btn"
                       class="ml-4 px-3 py-1 bg-blue-500 text-white rounded hover:bg-blue-600 text-xs font-semibold opacity-80 hover:opacity-100 transition"
                       title="Download today's Excel file with all extracted data">
                        📥 Download Today's Excel
                    </a>
                </div>
            </div>

            <!-- Messages Area -->
            <div class="flex-1 overflow-y-auto custom-scrollbar p-4 bg-gray-50" id="messages-container">
                <div id="messages" class="space-y-4">
                    <!-- Messages will be inserted here -->
                </div>
            </div>

            <!-- Input Area -->
            <div class="border-t border-gray-200 p-4 bg-white">
                <div class="flex items-end space-x-2">
                    <label class="cursor-pointer p-2 hover:bg-gray-100 rounded-lg transition">
                        <input type="file" id="image-input" accept="image/*,application/pdf,.doc,.docx" class="hidden">
                        <svg class="w-6 h-6 text-gray-600" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                            <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M15.172 7l-6.586 6.586a2 2 0 102.828 2.828l6.414-6.586a4 4 0 00-5.656-5.656l-6.415 6.585a6 6 0 108.486 8.486L20.5 13"></path>
                        </svg>
                    </label>
                    <div class="flex-1">
                        <textarea 
                            id="message-input" 
                            placeholder="Type a message..." 
                            class="w-full p-3 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 teams-purple-ring focus:border-transparent resize-none"
                            rows="1"
                            onkeydown="if(event.key === 'Enter' && !event.shiftKey) { event.preventDefault(); sendMessage(); }"
                        ></textarea>
                    </div>
                    <button 
                        id="send-button"
                        onclick="sendMessage()" 
                        class="teams-purple text-white px-6 py-3 rounded-lg font-semibold transition disabled:opacity-50 disabled:cursor-not-allowed"
                        style="background-color: #4B53BC;"
                        onmouseover="this.style.backgroundColor='#3d44a0'"
                        onmouseout="this.style.backgroundColor='#4B53BC'"
                    >
                        Send
                    </button>
                </div>
            </div>
        </div>
    </div>

    <script>
        // User identification
        let currentUserName = '';
        let currentUserId = '';
        let ws = null;
        const messagesContainer = document.getElementById('messages');
        const messageInput = document.getElementById('message-input');
        const sendButton = document.getElementById('send-button');
        const imageInput = document.getElementById('image-input');
        const messagesScrollContainer = document.getElementById('messages-container');
        let currentChat = 'Team Alpha';
        
        // Show user identification modal on page load
        function showUserModal() {
            const modal = document.createElement('div');
            modal.id = 'user-modal';
            modal.className = 'fixed inset-0 bg-black bg-opacity-50 flex items-center justify-center z-50';
            modal.innerHTML = `
                <div class="bg-white rounded-lg p-6 max-w-md w-full mx-4">
                    <h2 class="text-xl font-semibold mb-4 text-gray-800">Enter Your Information</h2>
                    <div class="space-y-4">
                        <div>
                            <label class="block text-sm font-medium text-gray-700 mb-1">Your Name *</label>
                            <input type="text" id="user-name-input" placeholder="e.g., John Doe" 
                                class="w-full p-2 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 teams-purple-ring"
                                onkeydown="if(event.key === 'Enter') connectWebSocket()">
                        </div>
                        <div>
                            <label class="block text-sm font-medium text-gray-700 mb-1">Your ID (Optional)</label>
                            <input type="text" id="user-id-input" placeholder="e.g., EMP001" 
                                class="w-full p-2 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 teams-purple-ring"
                                onkeydown="if(event.key === 'Enter') connectWebSocket()">
                        </div>
                        <button onclick="connectWebSocket()" 
                            class="w-full teams-purple text-white px-4 py-2 rounded-lg font-semibold hover:opacity-90"
                            style="background-color: #4B53BC;">
                            Join Chat
                        </button>
                    </div>
                </div>
            `;
            document.body.appendChild(modal);
        }
        
        // Connect WebSocket after user identification
        function connectWebSocket() {
            currentUserName = document.getElementById('user-name-input').value.trim();
            currentUserId = document.getElementById('user-id-input').value.trim();
            
            if (!currentUserName && !currentUserId) {
                alert('Please enter at least your name or ID');
                return;
            }
            
            // Remove modal
            const modal = document.getElementById('user-modal');
            if (modal) modal.remove();
            
            // Dynamically detect WebSocket protocol based on current page protocol
            const protocol = window.location.protocol === 'https:' ? 'wss:' : 'ws:';
            ws = new WebSocket(`${protocol}//${window.location.host}/ws`);
            
            // Setup WebSocket handlers
            setupWebSocketHandlers();
        }
        
        // Setup WebSocket event handlers
        function setupWebSocketHandlers() {
            // Auto-resize textarea
            messageInput.addEventListener('input', function() {
                this.style.height = 'auto';
                this.style.height = (this.scrollHeight) + 'px';
            });

            // Handle file/attachment selection - automatically triggers OCR
            imageInput.addEventListener('change', function(e) {
                const file = e.target.files[0];
                if (file) {
                    // Check if it's an image (OCR will be triggered automatically)
                    if (file.type.startsWith('image/')) {
                        console.log(`📎 Image file selected: ${file.name}, type: ${file.type}`);
                        const reader = new FileReader();
                        reader.onload = function(event) {
                            const base64Image = event.target.result;
                            // Send image - OCR will be triggered automatically on server
                            sendImageMessage(base64Image, file.name);
                        };
                        reader.readAsDataURL(file);
                    } else {
                        // For non-image files, show a message
                        alert('Currently only image files are supported for OCR. PDF and DOCX support coming soon.');
                    }
                    // Reset input
                    e.target.value = '';
                }
            });

            // Set status indicators (both are local now)
            document.getElementById('ocr-status-indicator').className = 'w-2 h-2 rounded-full bg-green-400';
            document.getElementById('extractor-status-indicator').className = 'w-2 h-2 rounded-full bg-green-400';

            // WebSocket event handlers
            ws.onopen = function() {
                console.log('WebSocket connected');
                sendButton.disabled = false;
                // Send user identification
                ws.send(JSON.stringify({
                    type: 'user_identify',
                    user_name: currentUserName,
                    user_id: currentUserId
                }));
            };

            ws.onmessage = function(event) {
                const data = JSON.parse(event.data);
                
                if (data.type === 'history') {
                    // Load chat history
                    messagesContainer.innerHTML = '';
                    data.messages.forEach(msg => {
                        addMessageToUI(msg);
                    });
                    scrollToBottom();
                } else if (data.type === 'message' || data.type === 'image') {
                    addMessageToUI(data);
                    scrollToBottom();
                } else if (data.type === 'notification') {
                    // Only show final notifications (not status updates)
                    addNotificationToUI(data);
                    scrollToBottom();
                }
            };

            ws.onerror = function(error) {
                console.error('WebSocket error:', error);
            };

            ws.onclose = function() {
                console.log('WebSocket disconnected');
                sendButton.disabled = true;
                // Attempt to reconnect after 3 seconds
                setTimeout(() => {
                    const protocol = window.location.protocol === 'https:' ? 'wss:' : 'ws:';
                    ws = new WebSocket(`${protocol}//${window.location.host}/ws`);
                    setupWebSocketHandlers();
                }, 3000);
            };
        }

        function addMessageToUI(data) {
            const messageDiv = document.createElement('div');
            messageDiv.className = 'flex items-start space-x-3';
            
            const avatar = document.createElement('div');
            avatar.className = 'w-8 h-8 rounded-full teams-purple flex items-center justify-center text-white font-semibold flex-shrink-0';
            avatar.textContent = data.sender ? data.sender.substring(0, 2).toUpperCase() : 'U';
            
            const contentDiv = document.createElement('div');
            contentDiv.className = 'flex-1';
            
            const senderName = document.createElement('div');
            senderName.className = 'text-sm font-semibold text-gray-800 mb-1';
            senderName.textContent = data.sender || 'Anonymous';
            
            if (data.type === 'image' && data.image) {
                const img = document.createElement('img');
                img.src = data.image;
                img.className = 'max-w-md rounded-lg shadow-md cursor-pointer';
                img.onclick = function() {
                    window.open(this.src, '_blank');
                };
                contentDiv.appendChild(senderName);
                contentDiv.appendChild(img);
                if (data.text) {
                    const textP = document.createElement('p');
                    textP.className = 'text-gray-700 mt-2';
                    textP.textContent = data.text;
                    contentDiv.appendChild(textP);
                }
            } else {
                const textP = document.createElement('p');
                textP.className = 'text-gray-700';
                textP.textContent = data.text || '';
                contentDiv.appendChild(senderName);
                contentDiv.appendChild(textP);
            }
            
            const timeSpan = document.createElement('span');
            timeSpan.className = 'text-xs text-gray-500 ml-2';
            timeSpan.textContent = data.timestamp || '';
            
            senderName.appendChild(timeSpan);
            
            messageDiv.appendChild(avatar);
            messageDiv.appendChild(contentDiv);
            messagesContainer.appendChild(messageDiv);
        }

        function sendMessage() {
            if (!ws || ws.readyState !== WebSocket.OPEN) {
                alert('Not connected. Please refresh and enter your information.');
                return;
            }
            
            const text = messageInput.value.trim();
            if (!text && !imageInput.files[0]) return;
            
            if (text) {
                const message = {
                    type: 'message',
                    text: text,
                    timestamp: new Date().toLocaleTimeString([], {hour: '2-digit', minute:'2-digit'})
                };
                
                ws.send(JSON.stringify(message));
                messageInput.value = '';
                messageInput.style.height = 'auto';
            }
        }

        function sendImageMessage(base64Image, filename) {
            if (!ws || ws.readyState !== WebSocket.OPEN) {
                alert('Not connected. Please refresh and enter your information.');
                return;
            }
            
            // Send image - OCR will be automatically triggered on server side
            console.log(`📎 Sending image: ${filename} - OCR will be triggered automatically`);
            const message = {
                type: 'image',
                image: base64Image,
                text: filename,
                timestamp: new Date().toLocaleTimeString([], {hour: '2-digit', minute:'2-digit'})
            };
            
            ws.send(JSON.stringify(message));
        }

        function addNotificationToUI(data) {
            const notificationDiv = document.createElement('div');
            notificationDiv.className = 'flex items-center justify-center my-2';
            
            const notificationContent = document.createElement('div');
            
            // Determine color based on status
            const status = data.status || 'success';
            let bgColor, borderColor, textColor;
            
            if (status === 'success') {
                bgColor = 'bg-green-100';
                borderColor = 'border-green-400';
                textColor = 'text-green-700';
            } else if (status === 'excel_fallback') {
                bgColor = 'bg-blue-100';
                borderColor = 'border-blue-400';
                textColor = 'text-blue-700';
            } else {
                bgColor = 'bg-red-100';
                borderColor = 'border-red-400';
                textColor = 'text-red-700';
            }
            
            notificationContent.className = `${bgColor} border ${borderColor} ${textColor} px-4 py-2 rounded-lg text-sm`;
            notificationContent.textContent = data.text || 'Completed';
            
            notificationDiv.appendChild(notificationContent);
            messagesContainer.appendChild(notificationDiv);
        }

        function scrollToBottom() {
            messagesScrollContainer.scrollTop = messagesScrollContainer.scrollHeight;
        }

        // Chat selection (for future enhancement)
        document.querySelectorAll('.chat-item').forEach(item => {
            item.addEventListener('click', function() {
                document.querySelectorAll('.chat-item').forEach(i => i.classList.remove('active-chat'));
                this.classList.add('active-chat');
                this.style.backgroundColor = '#f3f4f6';
                currentChat = this.dataset.chat;
                document.getElementById('chat-title').textContent = currentChat;
            });
        });
        
        // Show modal on page load
        window.addEventListener('DOMContentLoaded', function() {
            showUserModal();
        });
    </script>
</body>
</html>
    """
    return html_content

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)  # Initial connection without user info
    try:
        while True:
            data = await websocket.receive_text()
            message_data = json.loads(data)
            
            # Handle user identification
            if message_data.get('type') == 'user_identify':
                user_name = message_data.get('user_name')
                user_id = message_data.get('user_id')
                # Update user info for this connection
                if websocket in manager.active_connections:
                    if user_name:
                        manager.active_connections[websocket]['name'] = user_name
                    if user_id:
                        manager.active_connections[websocket]['id'] = user_id
                    if user_name or user_id:
                        manager.active_connections[websocket]['identifier'] = user_name or user_id
                continue
            
            # Get user info for this connection
            user_info = manager.get_user_info(websocket)
            message_data['sender'] = message_data.get('sender', user_info['identifier'])
            message_data['user_id'] = user_info.get('id', '')
            message_data['timestamp'] = message_data.get('timestamp', datetime.now().strftime('%H:%M'))
            
            # Process message through API if it's a new message (not history)
            if message_data.get('type') in ['message', 'image']:
                # Extract text and image
                text = message_data.get('text', '')
                image_base64 = message_data.get('image', None)
                
                # For image messages, don't send filename as text (only send the image)
                # For regular messages, send the text
                text_to_send = text if message_data.get('type') == 'message' else None
                
                # ALWAYS process if there's an image/attachment - trigger OCR automatically
                # Also process if there's text
                if image_base64 or text_to_send:
                    # Call API asynchronously (don't block message broadcast)
                    # OCR will be triggered automatically for any image
                    asyncio.create_task(process_and_save_message(
                        text_to_send, 
                        image_base64, 
                        message_data['timestamp'],
                        user_info['identifier']
                    ))
            
            # Broadcast message to all clients
            await manager.broadcast(message_data)
    except WebSocketDisconnect:
        manager.disconnect(websocket)

async def process_and_save_message(text: str, image_base64: Optional[str], timestamp: str, user_identifier: str):
    """
    Processes a message using local extractor (and OCR if image) and saves the result.
    Flow: Image → Local OCR → Extract text → Local Extractor → Google Sheets/Excel
    Runs silently in background - only shows final success/error notification.
    """
    try:
        final_text = None
        ocr_status = "skipped"
        
        # Step 1: ALWAYS trigger OCR for any image/attachment received
        # OCR is automatically triggered whenever an image/attachment is detected
        if image_base64:
            print("📷 Image/attachment detected - automatically triggering OCR...")
            ocr_text = await extract_text_from_image(image_base64)
            # Accept any non-empty text, even if very short
            if ocr_text and ocr_text.strip() and len(ocr_text.strip()) > 0:
                final_text = ocr_text.strip()
                ocr_status = "success"
                print(f"✓ OCR successful, extracted {len(final_text)} characters")
            else:
                ocr_status = "failed"
                print("✗ OCR failed or returned empty text - will retry with different settings")
        
        # Step 2: If we have text (from message or OCR), use it
        # If image OCR failed but we have original text, use that as fallback
        if not final_text and text:
            final_text = text.strip()
            print(f"Using provided text ({len(final_text)} characters)")
        
        # Step 3: Retry OCR if it failed (with different strategies)
        if not final_text or not final_text.strip():
            if image_base64:
                print("🔄 Retrying OCR with alternative strategies...")
                try:
                    # Retry with different PSM modes
                    if ',' in image_base64:
                        header, image_base64_clean = image_base64.split(',', 1)
                    else:
                        image_base64_clean = image_base64
                    
                    image_data = base64.b64decode(image_base64_clean)
                    image = Image.open(io.BytesIO(image_data))
                    
                    # Try different PSM (Page Segmentation Mode) configurations
                    psm_configs = [
                        ('', 'Default'),
                        ('--psm 6', 'Uniform block'),
                        ('--psm 11', 'Sparse text'),
                        ('--psm 12', 'Sparse text OSD'),
                        ('--psm 3', 'Fully automatic'),
                        ('--psm 4', 'Single column'),
                    ]
                    
                    for config, desc in psm_configs:
                        try:
                            text = pytesseract.image_to_string(image, config=config, lang='eng')
                            if text and text.strip() and len(text.strip()) > 0:
                                final_text = text.strip()
                                ocr_status = "success"
                                print(f"✓ OCR retry successful with {desc} (config: {config}), extracted {len(final_text)} characters")
                                break
                        except Exception as e:
                            print(f"  Retry with {desc} failed: {e}")
                            continue
                    
                    # If still no text, try with different image processing
                    if not final_text or not final_text.strip():
                        try:
                            # Try with higher contrast
                            img_gray = image.convert('L')
                            img_enhanced = ImageOps.autocontrast(img_gray, cutoff=2)
                            text = pytesseract.image_to_string(img_enhanced, lang='eng')
                            if text and text.strip() and len(text.strip()) > 0:
                                final_text = text.strip()
                                ocr_status = "success"
                                print(f"✓ OCR successful with enhanced contrast, extracted {len(final_text)} characters")
                        except Exception as e:
                            print(f"  Enhanced contrast retry failed: {e}")
                            
                except Exception as retry_error:
                    print(f"✗ OCR retry failed: {retry_error}")
                    import traceback
                    traceback.print_exc()
        
        # Final check - if still no text, show helpful error
        if not final_text or not final_text.strip():
            # Check if Tesseract is available
            try:
                pytesseract.get_tesseract_version()
                tesseract_available = True
            except:
                tesseract_available = False
            
            if not tesseract_available:
                error_msg = "✗ Tesseract OCR not found. Please install Tesseract:\n"
                error_msg += "  • macOS: brew install tesseract\n"
                error_msg += "  • Ubuntu: sudo apt-get install tesseract-ocr\n"
                error_msg += "  • Windows: Download from GitHub"
            else:
                error_msg = "✗ Could not extract text from image.\n"
                error_msg += "Please ensure:\n"
                error_msg += "  • Image is clear and readable\n"
                error_msg += "  • Text is not too small or blurry\n"
                error_msg += "  • Image contains visible text"
            
            error_notification = {
                "type": "notification",
                "text": error_msg,
                "status": "error",
                "timestamp": datetime.now().strftime('%H:%M')
            }
            await manager.broadcast(error_notification)
            print("✗ No text available for processing after all attempts")
            return
        
        print(f"🔍 Processing text with local extractor ({len(final_text)} characters)...")
        # Use local extractor on the OCR text or provided text
        result = await process_text_locally(final_text)
        
        if result.get('status') == 'success':
            extracted_data = result.get('extracted_data', {})
            
            # Add user identifier to extracted data
            extracted_data['user_identifier'] = user_identifier
            extracted_data['extracted_by'] = user_identifier
            extracted_data['extraction_timestamp'] = datetime.now().isoformat()
            
            # Create a filesystem-safe timestamp
            safe_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Save to Google Sheets or create downloadable Excel
            try:
                save_result = save_extracted_data(extracted_data, safe_timestamp)
                
                # Log success to console
                print(f"✓ Successfully processed message")
                if ocr_status == "success":
                    print(f"  (Processed via: Image → Local OCR → Local Extractor)")
                else:
                    print(f"  (Processed via: Text → Local Extractor)")
                
                # Handle different save results
                if save_result.get("status") == "success":
                    # Saved to Google Sheets
                    success_msg = f"✓ Data extracted and saved to Google Sheets: {save_result.get('message', '')}"
                    if ocr_status == "success":
                        success_msg += " (from image)"
                    
                    notification = {
                        "type": "notification",
                        "text": success_msg,
                        "status": "success",
                        "timestamp": datetime.now().strftime('%H:%M'),
                        "save_location": save_result.get('message')
                    }
                elif save_result.get("status") == "excel_fallback":
                    # Appended to daily Excel file
                    date_str = save_result.get('date', datetime.now().strftime("%Y-%m-%d"))
                    success_msg = f"✓ Data saved to daily Excel file ({date_str}). Use download button in header to get all data at end of day."
                    if ocr_status == "success":
                        success_msg += " (from image)"
                    
                    notification = {
                        "type": "notification",
                        "text": success_msg,
                        "status": "excel_fallback",
                        "timestamp": datetime.now().strftime('%H:%M')
                    }
                else:
                    # Unknown status
                    notification = {
                        "type": "notification",
                        "text": f"✓ Data extracted: {save_result.get('message', 'Saved')}",
                        "status": "success",
                        "timestamp": datetime.now().strftime('%H:%M')
                    }
                
                await manager.broadcast(notification)
            except Exception as save_error:
                error_msg = str(save_error)
                print(f"✗ Failed to save data: {error_msg}")
                import traceback
                traceback.print_exc()
                
                # Provide helpful error message
                user_error_msg = f"✗ Failed to save data: {error_msg[:150]}"
                
                error_notification = {
                    "type": "notification",
                    "text": user_error_msg,
                    "status": "error",
                    "timestamp": datetime.now().strftime('%H:%M')
                }
                await manager.broadcast(error_notification)
        else:
            error_msg = result.get('message', 'Unknown error')
            print(f"✗ Extraction failed: {error_msg}")
            
            error_notification = {
                "type": "notification",
                "text": f"✗ Extraction failed: {error_msg}",
                "status": "error",
                "timestamp": datetime.now().strftime('%H:%M')
            }
            await manager.broadcast(error_notification)
            
    except Exception as e:
        error_msg = str(e)
        print(f"✗ Error processing message: {error_msg}")
        
        error_notification = {
            "type": "notification",
            "text": f"✗ Error: {error_msg}",
            "status": "error",
            "timestamp": datetime.now().strftime('%H:%M')
        }
        await manager.broadcast(error_notification)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=10000)

